#encoding:utf-8
import openpyxl
import codecs


 
def copy_data(filePath):
    f=codecs.open(filePath,'rb','utf-8',errors='replace')
    wb = openpyxl.Workbook()
    sht=wb.create_sheet()
    for inDex,line in enumerate(f):
        l_orc=line.split('\t')
        for k in range(len(l_orc)):
            sht.cell(row=inDex+1,column=k+1).value=l_orc[k]
            
    print('write over')
    f.close()  
    print('start save')
    wb.save("C:\\Users\\Eric m shi\\Desktop\\111\\combine.xlsx")
    print('save ok')   
    
  
    
if __name__ == '__main__':
    #路径需修改，\用\\代替   
    rootdir="C:\\Users\\Eric m shi\\Desktop\\111\\GL1.txt"
    
    copy_data(rootdir)



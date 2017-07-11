'''
python版本 3.5.0

Author:Eric M Shi

运行前需要将excel中的数据去空格和不可打印字符，trim()和clean()

读取folder/data中" 1_CR样本-按笔.xlsx"的数据，写入"4_信贷审阅贷款模板.xlsx"中，按要求保存到output文件夹里面

请您根据CR字段示例和信贷审阅模版先编程。之后贷款库给到您，请您每个序号单独建excel 并以序号+支行名+客户名命名excel模版，
一个支行的所有模版放在一个文件夹中，以支行名命名文件夹。
'''

#encoding:utf-8

#所用到的库
import xlrd
import openpyxl
import os
import pandas as pd

if __name__ == '__main__':
    
    '''总体目录结构'''
    #data数据源的路径
    dataSourceFilePath=r'C:\Users\Eric m shi\Desktop\folder\data\1_CR样本-按笔.xlsx'
    
    #data模板的路径
    dataModelFilePath=r'C:\Users\Eric m shi\Desktop\folder\data\4_信贷审阅贷款模板.xlsx'
    
    #output的路径
    outputFilePath=r'C:\Users\Eric m shi\Desktop\folder\output'
    
    '''读数据'''
    #将excel数据读入到dataSourceDf中
    dataSourceDf=pd.read_excel(dataSourceFilePath)
    
    #data数据的列名
    colNameList=dataSourceDf.columns
    
    '''具体业务逻辑'''
    
    ###################按银行名称新建文件夹#############
    
    #银行名称列表
    bankNameList=dataSourceDf[colNameList[1]].drop_duplicates()
    
    
    #在output目录下按银行名称新建文件夹
    for bankName in bankNameList:
        folderName=os.path.join(outputFilePath,bankName)
        if not os.path.exists(folderName):
            os.makedirs(folderName)
    
    
    #复制模板到output文件夹
    for myIndex,group in dataSourceDf.groupby(colNameList[0]):
        newGroup=group.reset_index(drop=True)
        #写入B4:B7的数据
        dataRange1=list(newGroup[:1][colNameList[1:5]].values[0])
       
        #写入Row14:Row16和Row48的数据
        dataRange2=newGroup[colNameList[5:]].values
        
        #复制
        outputWb=openpyxl.load_workbook(dataModelFilePath)
        sht=outputWb.get_sheet_by_name('样本信息-客户经理填写')
        
        #复制dataRange1
        sht.cell(row=4,column=2).value=dataRange1[0]
        sht.cell(row=5,column=2).value=dataRange1[1]
        sht.cell(row=6,column=2).value=dataRange1[2]
        sht.cell(row=7,column=2).value=dataRange1[3]
         
        #复制dataRange2
        for theIndex,line in enumerate(dataRange2):
            for rowNum in range(13):
                if rowNum==0:
                    sht.cell(row=14+rowNum,column=2+theIndex).value=str(line[rowNum]).strip()
                else:
                    sht.cell(row=14+rowNum,column=2+theIndex).value=line[rowNum]
            sht.cell(row=48,column=2+theIndex).value=line[-1]    
         
        #将模板另存为   序号+支行名+客户名.xlsx
        outputWb.save(os.path.join(outputFilePath,dataRange1[0],str(myIndex)+'_'+dataRange1[0]+'_'+dataRange1[1]+".xlsx"))
                      
        

#encoding:utf-8
import openpyxl
import codecs
import os


#my_split：分隔符，my_code：编码的方式 
def copy_data(filePath,my_split,my_code):
    
    f=codecs.open(filePath,'rb',my_code,errors='ignore')
    try:
        wb = openpyxl.Workbook()
        sht=wb.active
#         for inDex,line in enumerate(f):
#             l_orc=line.split(my_split)[1:]
#             l_orc[11]=l_orc[11].replace(r'.','').replace(r',',r'.')
#             if l_orc[7]=='H':
#                     sht.cell(row=inDex+1,column=16).value='=0+'+l_orc[11]
#             if l_orc[7]=='S':
#                     sht.cell(row=inDex+1,column=15).value='=0+'+l_orc[11]
#             for k in range(len(l_orc)):
#                 sht.cell(row=inDex+1,column=k+1).value=l_orc[k]
            
        lines=f.readlines()
        for inDex,line in enumerate(lines[8:-2]):
            l_orc=line.split('\t')
            l_orc=l_orc[1:2]+l_orc[5:6]+l_orc[7:10]
            str_head=l_orc[0][0:8]
            l_orc.insert(0,str_head)
            for i in range(len(l_orc[2:6])):
                     
                    l_orc[2+i]=(l_orc[2+i].strip()).replace('.',"").replace(",",".")
                    if(l_orc[2+i].find('-')>0):
                        l_orc[2+i]=l_orc[2+i].replace("-",'')
                        l_orc[2+i]='-'+l_orc[2+i]
               
                
            for k in range(len(l_orc)):
                sht.cell(row=inDex+2,column=k+1).value=l_orc[k]
        sht.cell(row=1,column=1).value='AccountNo'
        sht.cell(row=1,column=2).value='.Item'
        sht.cell(row=1,column=3).value='FIRST PERIOD BAL'
        sht.cell(row=1,column=4).value='CURRENT DEBIT BAL'
        sht.cell(row=1,column=5).value='CURRENT CREDIT BAL'
        sht.cell(row=1,column=6).value='Current year balance'
        wb.save(filePath.replace('.txt','.xlsx'))        
    except Exception as e:
        print(filePath+':LineNum:'+str(inDex+1))
    finally:     
        f.close()  
   
    
def list_file(file_fun,rootdir,*args):
    list=os.listdir(rootdir)
     
    for i in range(0,len(list)):
        path=os.path.join(rootdir,list[i])
        if os.path.isfile(path):
            file_fun(path,*args)

if __name__ == '__main__':
    mydir=r"C:\Users\Eric m shi\Desktop\competition\TB_Trans"
    list_file(copy_data, mydir,'\t','utf-8') 
    
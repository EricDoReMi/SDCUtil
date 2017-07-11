#encoding:utf-8
import os
import openpyxl
import time
import codecs


 
def copy_data(fileName,filePath,rootdir,mysheet,jeSheetName):
    
    f=codecs.open(filePath,'r','utf-8')
    lines=f.readlines()
    l_orc=[]
    l_desc=[]
    l_desc.append(['AccountNo', '.Item', 'FIRST PERIOD BAL', 'CURRENT DEBIT BAL', 'CURRENT CREDIT BA', 'Current year balance','JE DEBIT SUM','JE CREDIT SUM','DIM DEBIT','DIM CREDIT','DIM'])
    for line in lines[8:-2]:
        l_orc=line.split('\t')
        l_orc=l_orc[1:2]+l_orc[5:6]+l_orc[7:]
        str_head=l_orc[0][0:8].strip()
        l_orc.insert(0,str_head)
        for i in range(len(l_orc[2:6])):
            l_orc[2+i]=(l_orc[2+i].strip()).replace('.',"").replace(",",".")
            if(l_orc[2+i].find('-')>0):
                l_orc[2+i]=l_orc[2+i].replace("-",'')
                l_orc[2+i]='-'+l_orc[2+i]
       
        
        l_desc.append(l_orc)
        
        for j in range(len(l_desc)):
            for k in range(len(l_desc[j])):
                mysheet.cell(row=j+1,column=k+1).value=l_desc[j][k]
            if not j:
                continue
            mysheet.cell(row=j+1,column=7).value="=VLOOKUP(A"+str(j+1)+","+jeSheetName+"!A:C,2)"
            mysheet.cell(row=j+1,column=8).value="=VLOOKUP(A"+str(j+1)+","+jeSheetName+"!A:C,3)"
            mysheet.cell(row=j+1,column=9).value="=D"+str(j+1)+"-G"+str(j+1)
            mysheet.cell(row=j+1,column=10).value="=E"+str(j+1)+"-I"+str(j+1)
            mysheet.cell(row=j+1,column=11).value="=I"+str(j+1)+"-J"+str(j+1)
            
    f.close()
  
    
if __name__ == '__main__':
    
    #获取当前时间
    my_year_month=time.strftime("%Y-%D")
    my_year_month=my_year_month[0:4]+"_"+my_year_month[5:7]+"_"+my_year_month[8:10]
     
       
    #路径需修改，\用\\代替   
    rootdir=r"C:\Users\Eric m shi\Desktop\GLPCATB2"
    
    listDir=os.listdir(rootdir)
    for m in range(len(listDir)):
        wb = openpyxl.Workbook() #新建一张excel表用于保存数据
        wbPath=rootdir+"\\"+listDir[m]
        list=os.listdir(wbPath)#每个工作表中的txt文件
           
        for i in range(len(list)):
            path=os.path.join(wbPath,list[i])
            fileName=os.path.basename(path)
            shtName=fileName.split('_')[-1].split('.')[-2].split('P')[1]
            if len(shtName)<2:
                shtName='0'+shtName
            mysheet=wb.create_sheet(title="TB"+shtName)
            myJESumSheet=wb.create_sheet(title="JE"+shtName)
            
            if os.path.isfile(path):
                copy_data(fileName,path,rootdir,mysheet,myJESumSheet.title)
        wb.remove_sheet(wb.get_sheet_by_name('Sheet')) 
        wb.save(wbPath+"\\"+listDir[m]+"_"+my_year_month+".xlsx")


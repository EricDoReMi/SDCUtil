#encoding:utf-8
import openpyxl
import codecs

if __name__ == '__main__':
    
    f=codecs.open(r'C:\Users\Eric m shi\Desktop\333\DcNo.csv','r','utf-8')
    lines=f.readlines()
    len_line=len(lines)
    l_desc=[]
    for inDex,line in enumerate(lines):
        l_orc=line.split(',')
        l_orc[1]=""
        if inDex+1<len_line and int(l_orc[0])+1!=int(lines[inDex+1].split(',')[0]):
                l_orc[1]='不连续'
                l_desc.append(l_orc)
    f.close()
      
    wb=openpyxl.Workbook()#新建立一个工作簿wb1  
    ws = None   
    for inDex,line in enumerate(l_desc):
        if inDex%1000000==0:
            ws = wb.create_sheet()
            print(ws)
            ws.cell(row=1,column=1).value='DocNo'
            ws.cell(row=1,column=2).value='是否是连续的'
            
        ws.cell(row=inDex%500000+2,column=1).value=line[0]
        ws.cell(row=inDex%500000+2,column=2).value=line[1]

    print('start save')     
    wb.save(r'C:\Users\Eric m shi\Desktop\333\凭证号不唯一.xlsx')
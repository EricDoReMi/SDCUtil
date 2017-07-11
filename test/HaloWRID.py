#encoding:utf-8

import openpyxl
'''
Created on 6 Mar 2017

@author: Eric M Shi
'''

def appendValue(v,l):
    if l is None:
        l=[]
    l.append(v)
    return l 

if __name__ == '__main__':
    wb=openpyxl.load_workbook(r"C:\Users\Eric m shi\Desktop\666\1.xlsx")
    sht=wb.get_sheet_by_name("data")
    
    #将元素写入字典中   
    myDic={}#用于存放key和listValue
    rowNum=2
    while True:
        if sht.cell(row=rowNum,column=1).value is None:break
        k=sht.cell(row=rowNum,column=1).value
        v=sht.cell(row=rowNum,column=2).value
        try:
            l=myDic[k]
        except KeyError:
            l=None
        myDic[k]=appendValue(v,l)
        rowNum+=1
    
    #将结果写入一个表中
    resultSht=wb.create_sheet('result', True)
    
    rowNum=1
    for key,value in myDic.items():
        resultSht.cell(row=rowNum,column=1).value=key
        resultSht.cell(row=rowNum,column=2).value='_'.join(value)
        rowNum+=1;
    
    wb.save(r"C:\Users\Eric m shi\Desktop\666\new.xlsx")
    
  
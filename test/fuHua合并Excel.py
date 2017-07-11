#encoding:utf-8

import openpyxl
import os
import sys

def copy_data(filePath,dataStore):
    wb=openpyxl.load_workbook(filePath)
    
    shts=wb.get_sheet_names()
    
    
    
    #循环遍历所有的sheet
    for i in range(len(shts)):
        #获得sheet
        sht=wb.get_sheet_by_name(shts[i])
        #每张sheet的起始行数
        rstart=1
        if shts[i]=='Sheet1':
            rstart=2
          
        for r in range(rstart,sht.max_row+1):
            if str(sht.cell(row=r,column=1).value) not in['None','本月合计','本年累计']:
                
                dataStore.append('\t'.join([str(sht.cell(row=r,column=c).value) for c in range(1,sht.max_column+1)]))



if __name__ == '__main__':
    #用于存放最终数据的二维list
    dataStore=[]
    dataStore.append('处理\t凭证号\t日期\t帐户\t部门\t产品\t往来\t预留\t名称\t参考\t货币\t汇率\t银行\t原币借方\t本币借方\t原币贷方\t本币贷方\t原币余额\t本币余额')
    
    rootdir="C:\\Users\\Eric m shi\\Desktop\\test"
    
    listNames=os.listdir(rootdir)
    for m in range(len(listNames)):
        wbPath=rootdir+"\\"+listNames[m]#excel文件的路径
        copy_data(wbPath,dataStore)
        
    f1=open('C:\\Users\\Eric m shi\\Desktop\\test\\combine.txt','w+')
          
    for j in dataStore:
        f1.writelines(j)
      
    f1.close()         
            
         
        
    